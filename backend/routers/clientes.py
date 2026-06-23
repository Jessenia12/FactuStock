from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import math, io, csv

from database import get_db
from models import PersonaComercial, UsuarioSistema, FacturaVenta
from schemas import (
    PersonaComercialCreate, PersonaComercialUpdate, PersonaComercialResponse
)
from dependencies import get_current_user

router = APIRouter(prefix="/api/clientes", tags=["Clientes"])


# ─── GET /api/clientes ────────────────────────────────────
@router.get("/", response_model=dict)
def listar_clientes(
    pagina: int = Query(1, ge=1),
    por_pagina: int = Query(20, ge=1, le=200),
    buscar: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    query = db.query(PersonaComercial).filter(
        PersonaComercial.id_usuario == current_user.id_usuario,
        PersonaComercial.flag_cliente == True
    )
    if buscar:
        query = query.filter(
            PersonaComercial.nombres_apellidos.contains(buscar) |
            PersonaComercial.razon_social.contains(buscar) |
            PersonaComercial.identificacion.contains(buscar) |
            PersonaComercial.email.contains(buscar)
        )
    total = query.count()
    clientes = (
        query.order_by(PersonaComercial.nombres_apellidos.asc())
        .offset((pagina - 1) * por_pagina)
        .limit(por_pagina)
        .all()
    )
    return {
        "items": [PersonaComercialResponse.from_orm(c) for c in clientes],
        "total": total,
        "pagina": pagina,
        "por_pagina": por_pagina,
        "total_paginas": math.ceil(total / por_pagina) if total > 0 else 1
    }


# ─── GET /api/clientes/exportar/{formato} ─────────────────
@router.get("/exportar/{formato}")
def exportar_clientes(
    formato: str,
    buscar: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    if formato not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Formato no valido. Use 'csv' o 'xlsx'.")

    query = db.query(PersonaComercial).filter(
        PersonaComercial.id_usuario == current_user.id_usuario,
        PersonaComercial.flag_cliente == True
    )
    if buscar:
        query = query.filter(
            PersonaComercial.nombres_apellidos.contains(buscar) |
            PersonaComercial.razon_social.contains(buscar) |
            PersonaComercial.identificacion.contains(buscar) |
            PersonaComercial.email.contains(buscar)
        )
    clientes = query.order_by(PersonaComercial.nombres_apellidos.asc()).all()

    # ─── Encabezados SOLO con los campos reales de la BD ───
    COL_LABELS = [
        "Identificación",
        "Tipo Identificación",
        "Nombres / Apellidos",
        "Razón Social",
        "Teléfono",
        "Email",
        "Dirección",
    ]

    def tipo_a_numero(tipo_str):
        """Convierte 'CEDULA' -> '1', 'RUC' -> '2', 'PASAPORTE' -> '3'. Nunca devuelve vacío."""
        if not tipo_str:
            return "1"   # Por defecto, asumir CÉDULA
        t = str(tipo_str).strip().upper()
        if t == "CEDULA":
            return "1"
        if t == "RUC":
            return "2"
        if t == "PASAPORTE":
            return "3"
        # Si ya es número (1,2,3) devolverlo como string
        if t in ("1", "2", "3"):
            return t
        # Cualquier otro valor -> CÉDULA por defecto
        return "1"

    def fila_datos(c):
        return [
            c.identificacion,
            tipo_a_numero(c.tipo_identificacion),  # siempre string "1", "2" o "3"
            c.nombres_apellidos or "",
            c.razon_social or "",
            c.telefono or "",
            c.email or "",
            c.direccion or "",
        ]

    # ── CSV ────────────────────────────────────────────────
    if formato == "csv":
        output = io.StringIO()
        # Usar QUOTE_NONNUMERIC para forzar comillas en el tipo (así Excel no lo interpreta mal)
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(COL_LABELS)
        for c in clientes:
            writer.writerow(fila_datos(c))
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(('\ufeff' + output.getvalue()).encode("utf-8")),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=clientes.csv",
                "Access-Control-Expose-Headers": "Content-Disposition",
            }
        )

    # ── XLSX (diseño limpio) ──────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl no instalado. Ejecuta: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Estilo del encabezado
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    header_fill = PatternFill("solid", fgColor="15389A")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D7E3"),
        right=Side(style="thin", color="D0D7E3"),
        top=Side(style="thin", color="D0D7E3"),
        bottom=Side(style="thin", color="D0D7E3")
    )

    ws.row_dimensions[1].height = 30
    for col_idx, label in enumerate(COL_LABELS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Estilo de datos
    data_font = Font(name="Calibri", size=10, color="1E293B")
    data_align = Alignment(vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    for row_idx, c in enumerate(clientes, 2):
        datos = fila_datos(c)
        ws.row_dimensions[row_idx].height = 18
        for col_idx, valor in enumerate(datos, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 2:  # Tipo Identificación -> centrado
                cell.alignment = center_align
            elif col_idx == 1:  # Identificación -> centrado y mono
                cell.alignment = center_align
                cell.font = Font(name="Courier New", size=10, color="334155")
            elif col_idx == 5:  # Teléfono -> centrado
                cell.alignment = center_align
            else:
                cell.alignment = data_align

    # Anchos de columna
    column_widths = [16, 14, 32, 28, 14, 30, 36]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COL_LABELS))}1"

    # Fila de totales
    total_row = len(clientes) + 2
    total_cell = ws.cell(row=total_row, column=1,
                         value=f"Total: {len(clientes)} cliente{'s' if len(clientes) != 1 else ''}")
    total_cell.font = Font(name="Calibri", size=9, bold=True, color="64748B", italic=True)
    total_cell.fill = PatternFill("solid", fgColor="F1F5F9")
    total_cell.border = thin_border
    total_cell.alignment = Alignment(horizontal="left", vertical="center")
    for col_idx in range(2, len(COL_LABELS) + 1):
        cell = ws.cell(row=total_row, column=col_idx, value="")
        cell.fill = PatternFill("solid", fgColor="F1F5F9")
        cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=clientes.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
    )


# ─── GET /api/clientes/{id} ───────────────────────────────
@router.get("/{id_cliente}", response_model=PersonaComercialResponse)
def obtener_cliente(
    id_cliente: int,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    cliente = db.query(PersonaComercial).filter(
        PersonaComercial.id_persona_comercial == id_cliente,
        PersonaComercial.id_usuario == current_user.id_usuario,
        PersonaComercial.flag_cliente == True
    ).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return cliente


# ─── POST /api/clientes ───────────────────────────────────
@router.post("/", response_model=PersonaComercialResponse, status_code=status.HTTP_201_CREATED)
def crear_cliente(
    datos: PersonaComercialCreate,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    existente = db.query(PersonaComercial).filter(
        PersonaComercial.id_usuario == current_user.id_usuario,
        PersonaComercial.identificacion == datos.identificacion
    ).first()

    if existente:
        if not existente.flag_cliente:
            existente.flag_cliente = True
            db.commit()
            db.refresh(existente)
            return existente
        raise HTTPException(
            status_code=400,
            detail=f"Ya existe un cliente con la identificacion {datos.identificacion}"
        )

    cliente = PersonaComercial(
        id_usuario=current_user.id_usuario,
        flag_cliente=True,
        flag_proveedor=getattr(datos, 'flag_proveedor', False),
        tipo_identificacion=datos.tipo_identificacion,
        identificacion=datos.identificacion,
        razon_social=datos.razon_social,
        nombres_apellidos=datos.nombres_apellidos,
        direccion=datos.direccion,
        telefono=datos.telefono,
        email=datos.email,
    )
    db.add(cliente)
    db.commit()
    db.refresh(cliente)
    return cliente


# ─── PUT /api/clientes/{id} ───────────────────────────────
@router.put("/{id_cliente}", response_model=PersonaComercialResponse)
def actualizar_cliente(
    id_cliente: int,
    datos: PersonaComercialUpdate,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    cliente = db.query(PersonaComercial).filter(
        PersonaComercial.id_persona_comercial == id_cliente,
        PersonaComercial.id_usuario == current_user.id_usuario,
        PersonaComercial.flag_cliente == True
    ).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    datos_dict = datos.dict(exclude_unset=True)

    nueva_id = datos_dict.get('identificacion')
    if nueva_id and nueva_id != cliente.identificacion:
        existente = db.query(PersonaComercial).filter(
            PersonaComercial.id_usuario == current_user.id_usuario,
            PersonaComercial.identificacion == nueva_id,
            PersonaComercial.id_persona_comercial != id_cliente
        ).first()
        if existente:
            raise HTTPException(status_code=400, detail="Ya existe un registro con esa identificacion")

    CAMPOS_EDITABLES = {
        "tipo_identificacion", "identificacion", "nombres_apellidos",
        "razon_social", "direccion", "telefono", "email",
    }
    for campo, valor in datos_dict.items():
        if campo in CAMPOS_EDITABLES:
            setattr(cliente, campo, valor)

    db.commit()
    db.refresh(cliente)
    return cliente


# ─── DELETE /api/clientes/{id} ────────────────────────────
@router.delete("/{id_cliente}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_cliente(
    id_cliente: int,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    cliente = db.query(PersonaComercial).filter(
        PersonaComercial.id_persona_comercial == id_cliente,
        PersonaComercial.id_usuario == current_user.id_usuario,
        PersonaComercial.flag_cliente == True
    ).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    tiene_facturas = db.query(FacturaVenta).filter(
        FacturaVenta.id_persona_comercial == id_cliente
    ).first()

    if tiene_facturas:
        cliente.flag_cliente = False
        db.commit()
        return

    if cliente.flag_proveedor:
        cliente.flag_cliente = False
        db.commit()
    else:
        db.delete(cliente)
        db.commit()