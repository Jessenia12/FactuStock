from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional, List
import math, io, csv

from database import get_db
from models import Producto, CategoriaProducto, UsuarioSistema, DetalleFacturaVenta
from schemas import ProductoCreate, ProductoUpdate, ProductoResponse
from dependencies import get_current_user
from pydantic import BaseModel

router = APIRouter(prefix="/api/productos", tags=["Productos"])


# ====================== MODELO PARA IMPORTACIÓN ======================
class ProductoImport(BaseModel):
    codigo: str
    nombre: str
    descripcion: Optional[str] = None
    precio_unitario: float
    porcentaje_iva: float
    stock: int = 0
    stock_minimo: int = 5
    id_categoria: Optional[int] = None


# ====================== EXPORTACIÓN (CSV y XLSX) ======================
@router.get("/exportar/{formato}")
def exportar_productos(
    formato: str,
    buscar: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    if formato not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Formato no valido. Use 'csv' o 'xlsx'.")

    query = db.query(Producto).filter(
        Producto.id_usuario == current_user.id_usuario,
        Producto.activo == True
    )
    if buscar:
        query = query.filter(
            Producto.nombre.contains(buscar) |
            Producto.codigo.contains(buscar) |
            Producto.descripcion.contains(buscar)
        )
    productos = query.order_by(Producto.nombre.asc()).all()

    COL_LABELS = [
        "Código", "Nombre", "Descripción", "Precio Unitario",
        "IVA %", "Stock", "Stock Mínimo", "Estado"
    ]

    def estado_str(p):
        if p.stock == 0:
            return "Sin stock"
        if p.stock <= p.stock_minimo:
            return "Stock bajo"
        return "En stock"

    def fila_datos(p):
        return [
            p.codigo,
            p.nombre,
            p.descripcion or "",
            f"{p.precio_unitario:.2f}",
            f"{p.porcentaje_iva:.0f}",
            p.stock,
            p.stock_minimo,
            estado_str(p),
        ]

    # CSV
    if formato == "csv":
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(COL_LABELS)
        for p in productos:
            writer.writerow(fila_datos(p))
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(('\ufeff' + output.getvalue()).encode("utf-8")),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=productos.csv",
                "Access-Control-Expose-Headers": "Content-Disposition",
            }
        )

    # XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl no instalado. Ejecuta: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Estilo encabezado
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    header_fill = PatternFill("solid", fgColor="15389A")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D7E3"),
        right=Side(style="thin", color="D0D7E3"),
        top=Side(style="thin", color="D0D7E3"),
        bottom=Side(style="thin", color="D0D7E3")
    )

    ws.row_dimensions[1].height = 30
    for col_idx, label in enumerate(COL_LABELS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Estilo datos
    data_font = Font(name="Calibri", size=10, color="1E293B")
    data_align = Alignment(vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    for row_idx, p in enumerate(productos, 2):
        datos = fila_datos(p)
        ws.row_dimensions[row_idx].height = 18
        for col_idx, valor in enumerate(datos, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in (1, 6, 7):      # Código, Stock, Stock Mínimo → centro
                cell.alignment = center_align
            elif col_idx == 4:            # Precio → derecha
                cell.alignment = right_align
            else:
                cell.alignment = data_align

    column_widths = [15, 35, 30, 14, 8, 10, 12, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COL_LABELS))}1"

    total_row = len(productos) + 2
    total_cell = ws.cell(row=total_row, column=1,
                         value=f"Total: {len(productos)} producto{'s' if len(productos) != 1 else ''}")
    total_cell.font = Font(name="Calibri", size=9, bold=True, color="64748B", italic=True)
    total_cell.fill = PatternFill("solid", fgColor="F1F5F9")
    total_cell.border = thin_border
    total_cell.alignment = Alignment(horizontal="left", vertical="center")
    for col_idx in range(2, len(COL_LABELS) + 1):
        cell = ws.cell(row=total_row, column=col_idx, value="")
        cell.fill = PatternFill("solid", fgColor="F1F5F9")
        cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=productos.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
    )


# ====================== IMPORTACIÓN MASIVA ======================
@router.post("/importar", status_code=status.HTTP_201_CREATED)
def importar_productos(
    productos_data: List[ProductoImport],
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    """
    Importa una lista de productos.
    Valida que no existan códigos duplicados y que las categorías pertenezcan al usuario.
    """
    resultados = {"creados": 0, "errores": []}

    for idx, p in enumerate(productos_data):
        # Verificar código duplicado
        existente = db.query(Producto).filter(
            Producto.id_usuario == current_user.id_usuario,
            Producto.codigo == p.codigo,
            Producto.activo == True
        ).first()
        if existente:
            resultados["errores"].append({
                "fila": idx + 1,
                "codigo": p.codigo,
                "error": f"El código '{p.codigo}' ya existe"
            })
            continue

        # Validar categoría si se proporciona
        if p.id_categoria is not None:
            categoria = db.query(CategoriaProducto).filter(
                CategoriaProducto.id_categoria == p.id_categoria,
                CategoriaProducto.id_usuario == current_user.id_usuario
            ).first()
            if not categoria:
                resultados["errores"].append({
                    "fila": idx + 1,
                    "codigo": p.codigo,
                    "error": f"Categoría con ID {p.id_categoria} no encontrada"
                })
                continue

        nuevo = Producto(
            id_usuario=current_user.id_usuario,
            codigo=p.codigo,
            nombre=p.nombre,
            descripcion=p.descripcion,
            precio_unitario=p.precio_unitario,
            porcentaje_iva=p.porcentaje_iva,
            stock=p.stock,
            stock_minimo=p.stock_minimo,
            id_categoria=p.id_categoria,
            activo=True,
            costo_promedio=0.00   # se puede actualizar después con movimientos de inventario
        )
        db.add(nuevo)
        resultados["creados"] += 1

    db.commit()
    return resultados


# ====================== LISTAR PRODUCTOS ======================
@router.get("/", response_model=dict)
def listar_productos(
    pagina: int = Query(1, ge=1),
    por_pagina: int = Query(20, ge=1, le=200),
    buscar: Optional[str] = None,
    solo_con_stock: bool = Query(False),
    id_categoria: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    query = (
        db.query(Producto)
        .options(joinedload(Producto.categoria))
        .filter(
            Producto.id_usuario == current_user.id_usuario,
            Producto.activo == True
        )
    )
    if buscar:
        query = query.filter(
            Producto.nombre.contains(buscar) |
            Producto.codigo.contains(buscar) |
            Producto.descripcion.contains(buscar)
        )
    if solo_con_stock:
        query = query.filter(Producto.stock > 0)
    if id_categoria:
        query = query.filter(Producto.id_categoria == id_categoria)

    total = query.count()
    productos = (
        query.order_by(Producto.nombre.asc())
        .offset((pagina - 1) * por_pagina)
        .limit(por_pagina)
        .all()
    )
    return {
        "items": [ProductoResponse.from_orm(p) for p in productos],
        "total": total,
        "pagina": pagina,
        "por_pagina": por_pagina,
        "total_paginas": math.ceil(total / por_pagina) if total > 0 else 1
    }


# ====================== OBTENER PRODUCTO ======================
@router.get("/{id_producto}", response_model=ProductoResponse)
def obtener_producto(
    id_producto: int,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    producto = (
        db.query(Producto)
        .options(joinedload(Producto.categoria))
        .filter(
            Producto.id_producto == id_producto,
            Producto.id_usuario == current_user.id_usuario,
            Producto.activo == True
        ).first()
    )
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return producto


# ====================== CREAR PRODUCTO ======================
@router.post("/", response_model=ProductoResponse, status_code=status.HTTP_201_CREATED)
def crear_producto(
    datos: ProductoCreate,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    existente = db.query(Producto).filter(
        Producto.id_usuario == current_user.id_usuario,
        Producto.codigo == datos.codigo,
        Producto.activo == True
    ).first()
    if existente:
        raise HTTPException(
            status_code=400,
            detail=f"Ya existe un producto con el código '{datos.codigo}'"
        )
    if datos.id_categoria:
        categoria = db.query(CategoriaProducto).filter(
            CategoriaProducto.id_categoria == datos.id_categoria,
            CategoriaProducto.id_usuario == current_user.id_usuario
        ).first()
        if not categoria:
            raise HTTPException(status_code=404, detail="Categoría no encontrada")

    producto = Producto(
        id_usuario=current_user.id_usuario,
        id_categoria=datos.id_categoria,
        codigo=datos.codigo,
        nombre=datos.nombre,
        descripcion=datos.descripcion,
        precio_unitario=datos.precio_unitario,
        porcentaje_iva=datos.porcentaje_iva if datos.porcentaje_iva is not None else 15.00,
        stock=datos.stock if datos.stock is not None else 0,
        stock_minimo=datos.stock_minimo if datos.stock_minimo is not None else 5,
        costo_promedio=datos.costo_promedio if datos.costo_promedio is not None else 0.00,
        activo=True,
    )
    db.add(producto)
    db.commit()
    db.refresh(producto)
    return producto


# ====================== ACTUALIZAR PRODUCTO ======================
@router.put("/{id_producto}", response_model=ProductoResponse)
def actualizar_producto(
    id_producto: int,
    datos: ProductoUpdate,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    producto = db.query(Producto).filter(
        Producto.id_producto == id_producto,
        Producto.id_usuario == current_user.id_usuario,
        Producto.activo == True
    ).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    if datos.codigo and datos.codigo != producto.codigo:
        existente = db.query(Producto).filter(
            Producto.id_usuario == current_user.id_usuario,
            Producto.codigo == datos.codigo,
            Producto.id_producto != id_producto,
            Producto.activo == True
        ).first()
        if existente:
            raise HTTPException(status_code=400, detail="Ya existe un producto con ese código")

    for campo, valor in datos.dict(exclude_unset=True).items():
        setattr(producto, campo, valor)

    db.commit()
    db.refresh(producto)
    return producto


# ====================== ELIMINAR PRODUCTO ======================
@router.delete("/{id_producto}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_producto(
    id_producto: int,
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    producto = db.query(Producto).filter(
        Producto.id_producto == id_producto,
        Producto.id_usuario == current_user.id_usuario,
        Producto.activo == True
    ).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    en_facturas = db.query(DetalleFacturaVenta).filter(
        DetalleFacturaVenta.id_producto == id_producto
    ).first()

    if en_facturas:
        producto.activo = False
        db.commit()
    else:
        db.delete(producto)
        db.commit()


# ====================== LISTAR CATEGORÍAS ======================
@router.get("/categorias/lista", response_model=list)
def listar_categorias(
    db: Session = Depends(get_db),
    current_user: UsuarioSistema = Depends(get_current_user)
):
    categorias = db.query(CategoriaProducto).filter(
        CategoriaProducto.id_usuario == current_user.id_usuario
    ).order_by(CategoriaProducto.nombre.asc()).all()
    return [{"id_categoria": c.id_categoria, "nombre": c.nombre} for c in categorias]